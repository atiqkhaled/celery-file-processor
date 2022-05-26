
import sys
sys.path.append("./celery_tasks")
sys.path.append("..")
from cmath import infj
import logging
from fileinput import filename
from helper.property_reader import PropertyReader
from repo.mapper_repo import MapperRepo
from repo.file_content_repo import FileRepo
from repo.status_repo import StatusRepo
from helper.db_helper import DBHelper
import os
import openpyxl
from openpyxl import load_workbook
import pymongo
from bson.json_util import dumps, loads
from bson.objectid import ObjectId
import json
from app_worker import app
from celery import Task
from celery.exceptions import MaxRetriesExceededError
from azure.storage.blob import BlobServiceClient


sys.path.insert(0, os.path.realpath(os.path.pardir))

propertyReader = PropertyReader()
dbHelper = DBHelper()
fileRepo = FileRepo()
statusRepo = StatusRepo()


STORAGEACCOUNTURL = "https://<STORAGE_ACCOUNT_NAME>.blob.core.windows.net"
CONNECT_STR = "DefaultEndpointsProtocol=https;AccountName=azurestorageistiaq;AccountKey=ldjvW6l0LqCBEirGTr4KTCZfpFmcF7iF+ljGV50anqIyKZk9lk2j8N6ainx8KOJ/yM/sIvTaV2OD+AStWu8K2A==;EndpointSuffix=core.windows.net"

blob_service_client = BlobServiceClient.from_connection_string(CONNECT_STR)


def getMapping(mapper):
    return json.loads(mapper["modelContent"])


def getExcelHeader(rows):
    return [cell.value for cell in next(rows)]


def getDownloadedWorkbook(filePath):
    blob_client = blob_service_client.get_blob_client(
        container=propertyReader.getContainer(), blob=filePath)
    download_file_path = os.path.join("temp", filePath)

    with open(download_file_path, "wb") as download_file:
        download_file.write(blob_client.download_blob().readall())

    return load_workbook(download_file_path)


@app.task(ignore_result=False, bind=True)
def save_entry(self, fileId):
    try:
        fileContent = fileRepo.getFileContentById(fileId)
        filePath = fileContent["path"]
        fileName = fileContent["path"]
        mapperRepo = MapperRepo()
        mapper = mapperRepo.getMapperById(fileContent["mapperId"])
        mappingInfo = getMapping(mapper)
        collections = dbHelper.getCollection(mapper["tableName"])
        mappedHeaders = list(mappingInfo.values())
        # if propertyReader.isLocal:
        #     print("Local Bro")
        #     filePath = propertyReader.getUploadedPath() + "/" + fileName
        #     print(filePath)
        book = getDownloadedWorkbook(filePath)
        sheets = book.sheetnames
        for sheet in sheets:
            worksheet = book[sheet]
            rows = worksheet.rows
            headers = getExcelHeader(rows)
            headersDict = {k: v for v, k in enumerate(headers)}
            filterdHeaders = [
                head for head in headers if head in mappedHeaders]
            dbEntries = []
            for row in rows:
                dbEntry = {}
                for entry in mappingInfo:
                    excelColumn = mappingInfo[entry]
                    if excelColumn in filterdHeaders:
                        dbEntry[entry] = row[headersDict[excelColumn]].value
                    else:
                        dbEntry[entry] = ""
                dbEntries.append(dbEntry)

        collections.insert_many(dbEntries)
        updateStatus(fileId)
        return {'status': 'SUCCESS', 'result': 'done'}

    except Exception as ex:
        try:
            print("Exception")
            logging.info(ex)
            self.retry(countdown=1)
        except MaxRetriesExceededError as ex:
            return {'status': 'FAIL', 'result': 'max retried achieved'}


def updateStatus(fileId):
    injectedStatus = statusRepo.getStatusByName("INJECTED")
    statusIdtoUpdate = injectedStatus["_id"]
    fileContent = fileRepo.getFileContentById(fileId)
    currentStatusId = fileContent["status"]
    fileRepo.update(currentStatusId, statusIdtoUpdate)
